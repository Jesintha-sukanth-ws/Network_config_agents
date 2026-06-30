import io
import logging
import mimetypes
import re
import threading
import time
import json
import os
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
import requests
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth

from config.settings import (
    SERVICENOW_INSTANCE,
    SERVICENOW_USERNAME,
    SERVICENOW_PASSWORD,
    NETWORK_GROUP_ID,
    SERVICENOW_SSL_VERIFY,
    SERVICENOW_TIMEOUT,
    POLL_INTERVAL,
)

from app.services.orchestrator_service import process_task

# ----------------------------------------
# Logging
# ----------------------------------------

logger = logging.getLogger(__name__)

EXCEL_FILE_EXTENSIONS = {".xls", ".xlsx", ".csv"}
EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "application/octet-stream",
}

COLUMN_ALIASES = {
    "device_name": {"device_name", "hostname", "name", "device"},
    "vendor": {"vendor", "manufacturer", "make"},
    "os_type": {"os_type", "os", "operating_system", "platform"},
    "model_number": {"model_number", "model", "model_name", "product"},
    "management_host": {"management_host", "mgmt_host", "management_ip", "ip_address", "fqdn", "host"},
    "credentials.username": {"username", "user", "login"},
    "credentials.password": {"password", "pass", "secret"},
}


def _normalize_column_name(column_name: Any) -> str:
    if column_name is None:
        return ""
    normalized = str(column_name).strip().lower()
    normalized = normalized.replace(" ", "_").replace("-", "_")
    return normalized


def _safe_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_os_type(os_string: str) -> str:
    """
    Normalize OS type string to match device_capabilities.json keys.
    
    Handles various formats:
    - "cisco IOS-XE2026" -> "IOS-XE"
    - "IOS-XE 17.15" -> "IOS-XE"
    - "Cisco NX-OS 10.3" -> "NX-OS"
    - "NX-OS" -> "NX-OS"
    - "ios-xe" -> "IOS-XE"
    
    Args:
        os_string: Raw OS identifier from device data
        
    Returns:
        Normalized OS type matching device_capabilities.json
    """
    import re
    
    if not os_string:
        return ""
    
    original_string = os_string
    os_string = os_string.strip().upper()
    
    # Remove vendor prefixes (Cisco, Juniper, Arista, etc.)
    prefixes = ["CISCO ", "JUNIPER ", "ARISTA ", "NOKIA ", "DELL ", "HP ", "HUAWEI "]
    for prefix in prefixes:
        if os_string.startswith(prefix):
            os_string = os_string[len(prefix):].strip()
            break
    
    # Extract core OS type (everything before version numbers and special chars)
    # Handle cases like "IOS-XE2026", "IOS-XE 17.15", "NX-OS 10.3", etc.
    
    # Match known OS patterns
    patterns = [
        r"^(IOS-XE)",  # IOS-XE variations
        r"^(NX-OS)",   # NX-OS variations
        r"^(JUNOS)",   # Juniper JUNOS
        r"^(EOS)",     # Arista EOS
        r"^(IOS)(?![A-Z])",  # IOS (but not IOS-XE)
    ]
    
    for pattern in patterns:
        match = re.match(pattern, os_string)
        if match:
            normalized = match.group(1)
            logger.debug(f"Normalized os_type '{original_string}' -> '{normalized}'")
            return normalized
    
    # Fallback: return the first word/hyphenated group
    # e.g., "NX-OS 10.3" -> "NX-OS", "IOS-XE2026" -> "IOS-XE"
    match = re.match(r"^([A-Z0-9]+-?[A-Z0-9]*)", os_string)
    if match:
        normalized = match.group(1)
        logger.debug(f"Normalized os_type '{original_string}' -> '{normalized}' (fallback)")
        return normalized
    
    logger.debug(f"Could not normalize os_type '{original_string}', returning as-is")
    return os_string


def _map_excel_row_to_device_data(row: Dict[str, Any]) -> Dict[str, Any]:
    normalized = {
        _normalize_column_name(key): _safe_value(value)
        for key, value in row.items()
        if key is not None
    }

    mapped: Dict[str, Any] = {}
    credentials: Dict[str, str] = {}

    for destination, aliases in COLUMN_ALIASES.items():
        for source_name, value in normalized.items():
            if source_name in aliases and value:
                if destination.startswith("credentials."):
                    credentials[destination.split(".")[1]] = value
                else:
                    mapped[destination] = value
                break

    if credentials:
        mapped["credentials"] = {
            "username": credentials.get("username", ""),
            "password": credentials.get("password", ""),
        }

    if "model_number" not in mapped and normalized.get("model"):
        mapped["model_number"] = normalized.get("model")

    # Normalize os_type to match device_capabilities.json keys
    if "os_type" in mapped:
        mapped["os_type"] = _normalize_os_type(mapped["os_type"])

    return mapped


def _read_spreadsheet_rows(content: bytes, filename: str) -> List[List[Any]]:
    extension = Path(filename or "").suffix.lower()

    if extension == ".csv":
        df = pd.read_csv(io.StringIO(content.decode("utf-8", errors="replace")), dtype=str)
    else:
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
            sheet = workbook.active
            if sheet is None:
                raise ValueError("Spreadsheet contains no active worksheet")
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        except Exception:
            df = pd.read_excel(io.BytesIO(content), sheet_name=0, dtype=str)

    if df.empty:
        return []

    rows: List[List[Any]] = [list(df.columns)]
    rows.extend(df.fillna("").values.tolist())
    return rows


def _parse_device_data_from_excel(content: bytes, filename: str = "") -> Dict[str, Any]:
    rows = _read_spreadsheet_rows(content, filename)
    if not rows or len(rows) < 2:
        raise ValueError("Excel attachment contains no usable rows")

    headers = [_normalize_column_name(value) for value in rows[0]]
    if not any(headers):
        raise ValueError("Excel attachment header row is invalid")

    candidate_rows: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_map = {
            headers[index]: row[index]
            for index in range(min(len(headers), len(row)))
            if headers[index]
        }
        candidate_rows.append(row_map)

    if not candidate_rows:
        raise ValueError("Excel attachment contains no data rows")

    for row_data in candidate_rows:
        mapped = _map_excel_row_to_device_data(row_data)
        if mapped.get("vendor") and mapped.get("os_type") and mapped.get("management_host") and (mapped.get("model_number") or mapped.get("device_name")):
            return mapped

    return _map_excel_row_to_device_data(candidate_rows[0])


def _attachment_is_excel(attachment: Dict[str, Any]) -> bool:
    file_name = _safe_value(attachment.get("file_name", ""))
    content_type = _safe_value(attachment.get("content_type", ""))
    extension = Path(file_name).suffix.lower()
    if extension in EXCEL_FILE_EXTENSIONS:
        return True
    if content_type.lower() in EXCEL_MIME_TYPES:
        return True
    guessed_type, _ = mimetypes.guess_type(file_name)
    return guessed_type in EXCEL_MIME_TYPES


# ----------------------------------------
# ServiceNow Task States
# ----------------------------------------

TASK_STATES = {
    "OPEN": "1",
    "WORK_IN_PROGRESS": "2",
    "CLOSED_COMPLETE": "3",
    "CLOSED_INCOMPLETE": "4",
}


# ----------------------------------------
# Fetch next task
# ----------------------------------------

def get_next_open_task():
    """
    Fetch next open task from ServiceNow.
    
    Query criteria (AND conditions):
    - assignment_group = NETWORK_GROUP_ID (assigned to network team)
    - state = 1 (Open)
    - approval = "approved" (approved by manager)
    
    To test without matching tasks, set TEST_MODE_QUERY=true in .env
    """

    url = (
        f"{SERVICENOW_INSTANCE}"
        "/api/now/table/sc_task"
    )

    # PRODUCTION QUERY - Only fetch approved tasks assigned to network group
    production_query = (
        f"assignment_group={NETWORK_GROUP_ID}"
        "^state=1"
        "^approval=approved"
    )
    
    # TEST QUERY - Fetch ANY open state=1 tasks (even if not assigned/approved)
    # Use this to verify ServiceNow connectivity when no production tasks match
    test_query = "state=1"

    # Check if test mode is enabled
    test_mode = os.environ.get("TEST_MODE_QUERY", "false").lower() == "true"
    query_string = test_query if test_mode else production_query

    params = {
        "sysparm_query": query_string,
        "sysparm_limit": "1",
        "sysparm_fields": (
            "sys_id,"
            "number,"
            "short_description,"
            "description,"
            "variables,"
            "assignment_group,"
            "state,"
            "approval"
        ),
    }

    try:
        auth = HTTPBasicAuth(
            str(SERVICENOW_USERNAME or ""),
            str(SERVICENOW_PASSWORD or ""),
        )
        timeout_value = None
        if SERVICENOW_TIMEOUT is not None:
            try:
                timeout_value = float(SERVICENOW_TIMEOUT)
            except (TypeError, ValueError):
                timeout_value = None

        logger.info(f"ServiceNow query: {query_string} (test_mode={test_mode})")
        
        response = requests.get(
            url,
            auth=auth,
            params=params,
            verify=SERVICENOW_SSL_VERIFY,
            timeout=timeout_value,
        )
        response.raise_for_status()

        tasks = response.json().get("result", [])
        
        if tasks:
            task = tasks[0]
            logger.info(f"✓ Found task: {task.get('number')} (state={task.get('state')}, approval={task.get('approval')})")
            
            # Validate task meets production criteria if not in test mode
            if not test_mode:
                if task.get('state') != '1':
                    logger.warning(f"  ⚠ Task state is '{task.get('state')}', expected '1' (Open)")
                if task.get('approval') != 'approved':
                    logger.warning(f"  ⚠ Task approval is '{task.get('approval')}', expected 'approved'")
        else:
            logger.debug(f"✗ No matching tasks found. Test mode: {test_mode}")
            if not test_mode:
                logger.info("  💡 To troubleshoot, set TEST_MODE_QUERY=true in .env to see ALL open tasks")
        
        return tasks[0] if tasks else None

    except requests.RequestException as e:
        logger.error(f"✗ ServiceNow API error: {e}")
        return None


# ----------------------------------------
# Update task state
# ----------------------------------------

def update_task_state(task_sys_id: str, state: str) -> bool:

    url = (
        f"{SERVICENOW_INSTANCE}"
        f"/api/now/table/sc_task/{task_sys_id}"
    )

    payload = {"state": state}

    try:
        auth = HTTPBasicAuth(
            str(SERVICENOW_USERNAME or ""),
            str(SERVICENOW_PASSWORD or ""),
        )
        timeout_value = None
        if SERVICENOW_TIMEOUT is not None:
            try:
                timeout_value = float(SERVICENOW_TIMEOUT)
            except (TypeError, ValueError):
                timeout_value = None

        response = requests.patch(
            url,
            auth=auth,
            json=payload,
            verify=SERVICENOW_SSL_VERIFY,
            timeout=timeout_value,
        )
        if response.status_code != 200:
            logger.error(f"✗ State update failed. Status: {response.status_code}, Response: {response.text}")
            return False
        response.raise_for_status()
        logger.info(f"✓ Task state updated to {state}")
        return True

    except requests.RequestException as e:
        logger.error(f"✗ Failed updating task state: {e}")
        return False


# ----------------------------------------
# Update task work notes
# ----------------------------------------

def update_task_work_notes(task_sys_id: str, work_notes: str) -> bool:

    url = (
        f"{SERVICENOW_INSTANCE}"
        f"/api/now/table/sc_task/{task_sys_id}"
    )

    payload = {"work_notes": work_notes}

    try:
        logger.info(
            "Updating work notes for sc_task (%s)",
            task_sys_id,
        )
        auth = HTTPBasicAuth(
            str(SERVICENOW_USERNAME or ""),
            str(SERVICENOW_PASSWORD or ""),
        )
        timeout_value = None
        if SERVICENOW_TIMEOUT is not None:
            try:
                timeout_value = float(SERVICENOW_TIMEOUT)
            except (TypeError, ValueError):
                timeout_value = None

        response = requests.patch(
            url,
            auth=auth,
            json=payload,
            verify=SERVICENOW_SSL_VERIFY,
            timeout=timeout_value,
        )
        if response.status_code == 200:
            logger.info(
                "Work notes updated successfully for sc_task (%s)",
                task_sys_id,
            )
            return True

        logger.error(
            "Work note update failed. status=%s response=%s",
            response.status_code,
            response.text,
        )
        return False

    except requests.RequestException as e:
        logger.exception("Exception occurred while updating work notes")
        return False

    # ----------------------------------------
    # Get task attachments
    # ----------------------------------------

def get_task_attachments(task_sys_id: str) -> List[Dict[str, Any]]:

    url = (
        f"{SERVICENOW_INSTANCE}"
        "/api/now/table/sys_attachment"
    )

    params = {
        "sysparm_query": (
            f"table_name=sc_task"
            f"^table_sys_id={task_sys_id}"
        ),
        "sysparm_fields": (
            "sys_id,"
            "file_name,"
            "content_type"
        ),
    }

    try:
        auth = HTTPBasicAuth(
            str(SERVICENOW_USERNAME or ""),
            str(SERVICENOW_PASSWORD or ""),
        )
        timeout_value = None
        if SERVICENOW_TIMEOUT is not None:
            try:
                timeout_value = float(SERVICENOW_TIMEOUT)
            except (TypeError, ValueError):
                timeout_value = None

        response = requests.get(
            url,
            auth=auth,
            params=params,
            verify=SERVICENOW_SSL_VERIFY,
            timeout=timeout_value,
        )
        response.raise_for_status()
        return response.json().get("result", [])

    except requests.RequestException as e:
        logger.exception("Failed to fetch task attachments: %s", e)
        return []


# ----------------------------------------
# Download attachment
# ----------------------------------------

def download_attachment(attachment_sys_id: str) -> bytes:

    url = (
        f"{SERVICENOW_INSTANCE}"
        f"/api/now/attachment/{attachment_sys_id}/file"
    )

    auth = HTTPBasicAuth(
        str(SERVICENOW_USERNAME or ""),
        str(SERVICENOW_PASSWORD or ""),
    )
    timeout_value = None
    if SERVICENOW_TIMEOUT is not None:
        try:
            timeout_value = float(SERVICENOW_TIMEOUT)
        except (TypeError, ValueError):
            timeout_value = None

    response = requests.get(
        url,
        auth=auth,
        verify=SERVICENOW_SSL_VERIFY,
        timeout=timeout_value,
    )

    response.raise_for_status()
    return response.content


# ----------------------------------------
# Excel -> JSON
# ----------------------------------------

def excel_to_json(excel_bytes: bytes):
    sheets = pd.read_excel(
        io.BytesIO(excel_bytes),
        sheet_name=None,
        dtype=str,
    )

    result: Dict[str, Any] = {}

    for sheet_name, df in sheets.items():
        df = df.fillna("")
        result[sheet_name] = df.to_dict(orient="records")

    return result


# ----------------------------------------
# Enrich task with attachment data
# ----------------------------------------

def enrich_task_with_excel(task: Dict[str, Any]) -> Dict[str, Any]:
    task_id = task.get("sys_id")
    if not task_id:
        return task

    attachments = get_task_attachments(task_id)
    task["task_attachments"] = attachments

    for attachment in attachments:
        if not _attachment_is_excel(attachment):
            continue

        try:
            attachment_bytes = download_attachment(attachment["sys_id"])
            device_data = _parse_device_data_from_excel(
                attachment_bytes,
                attachment.get("file_name", ""),
            )
            if device_data:
                task["device_data"] = device_data
                logger.info(
                    "Loaded device details from attachment %s for task %s",
                    attachment.get("file_name"),
                    task_id,
                )
                return task
        except Exception as e:
            logger.exception(
                "Failed to parse Excel attachment %s for task %s: %s",
                attachment.get("file_name", "unknown"),
                task_id,
                e,
            )

    logger.warning(
        "No valid Excel attachment with device details found for task %s",
        task_id,
    )
    return task
# ----------------------------------------
# Poller Loop
# ----------------------------------------

def poll_servicenow(stop_event: threading.Event | None = None):
    """
    Long-running ServiceNow poller.
    
    Polls ServiceNow for approved tasks and processes them using the
    CR Lifecycle Agent framework. Safely falls back to legacy synchronous 
    orchestration if the lifecycle agent configuration is unavailable.
    
    Args:
        stop_event: Optional threading.Event to signal shutdown
    """

    logger.info("ServiceNow poller started")

    # Safe dynamic lookup of the lifecycle component to preserve backward compatibility
    try:
        from app.services.cr_lifecycle_agent import CRLifecycleAgent
        lifecycle_agent = CRLifecycleAgent()
        logger.info("Asynchronous Change Management lifecycle tracking enabled.")
    except (ImportError, ModuleNotFoundError, Exception) as err:
        lifecycle_agent = None
        logger.warning(
            "CRLifecycleAgent could not be loaded; falling back to legacy synchronous execution loop. Exception: %s", 
            err
        )

    # Initialize narrative service
    try:
        from app.services.itsm_narrative_service import (
            ITSMNarrativeService,
            NarrativeType,
        )
        narrative_service = ITSMNarrativeService()
    except Exception as e:
        narrative_service = None
        logger.warning("ITSMNarrativeService could not be initialized: %s", e)

    while True:

        if stop_event is not None and stop_event.is_set():
            logger.info("Polling stop requested; exiting loop")
            return

        work_done = False

        try:
            # ── PASS 1: Reconcile Outstanding, Long-running Approved Changes ────
            if lifecycle_agent is not None:
                try:
                    # Check active states tracked inside cr_tracking.json and drive executions
                    lifecycle_agent.reconcile_active_changes()
                except Exception as lifecycle_err:
                    logger.exception("Error during lifecycle tracking reconciliation: %s", lifecycle_err)

            # ── PASS 2: Ingest Brand New Incoming Request Contexts ──────────────
            task = get_next_open_task()

            if not task:
                logger.debug("No approved tasks found")
                if not work_done:
                    _wait(stop_event, int(POLL_INTERVAL) if POLL_INTERVAL is not None else 60)
                continue

            
            task_id = task.get("sys_id")
            task_number = task.get("number")

            # Defensive Check: Ensure this target isn't already caught in the change loop
            if (
                lifecycle_agent is not None
                and lifecycle_agent.is_tracked(task_number)
            ):
                logger.debug(
                    "Task %s is already explicitly tracked under a lifecycle workflow.",
                    task_number,
                )
                continue

            acquired = update_task_state(
                task_id,
                TASK_STATES["WORK_IN_PROGRESS"],
            )

            if not acquired:
                logger.warning("Unable to acquire %s", task_number)
                continue

            logger.info("✓ Acquired %s - now processing", task_number)
            work_done = True

            # Parse Excel attachments
            try:

                task = enrich_task_with_excel(task)

                logger.info(
                    "Loaded %s Excel attachment(s) for %s",
                    len(
                        task.get(
                            "task_attachments",
                            [],
                        )
                    ),
                    task_number,
                )

            except Exception as e:

                logger.exception(
                    "Failed to enrich task with Excel data: %s",
                    e,
                )

            # ── PASS 3: Route Execution Flow Based on System Topology Mode ──────
            if lifecycle_agent is not None:
                try:
                    # Route task through the CR lifecycle engine
                    result = lifecycle_agent.initialize_lifecycle(task)

                    if isinstance(result, dict) and result.get("status") == "failed":
                        logger.warning("%s lifecycle execution failed: %s", task_number, result)
                        
                        # Track this failed task for dashboard visibility
                        failure_reason = result.get("error_detail", result.get("error", "Validation failed"))
                        _track_failed_task(task_id, task_number, task, failure_reason)
                        
                        # Generate audit trail for failure
                        if narrative_service is not None:
                            try:
                                notes = narrative_service.generate_task_work_notes(
                                    narrative_type=NarrativeType.TASK_FAILURE,
                                    task_number=task_number,
                                    short_description=task.get("short_description", ""),
                                    description=task.get("description", ""),
                                    technical_details=result.get(
                                        "error_detail",
                                        result.get(
                                            "error",
                                            "Lifecycle initialization failed."
                                        )
                                    ),
                                )
                                logger.info(
                                    "Diagnostic: updating work notes for %s with generated content: %s",
                                    task_number,
                                    notes,
                                )
                                update_task_work_notes(task_id, notes)
                            except Exception as n_err:
                                logger.error(
                                    "Failed to generate/update failure narrative for %s: %s",
                                    task_number,
                                    n_err,
                                )

                        update_task_state(
                            task_id,
                            TASK_STATES["CLOSED_INCOMPLETE"],
                        )

                except Exception as e:
                    logger.exception(
                        "Failed to process lifecycle workflow for %s",
                        task_number,
                    )
                    
                    # Track this failed task for dashboard visibility
                    _track_failed_task(task_id, task_number, task, str(e))
                    
                    if narrative_service is not None:
                        try:
                            notes = narrative_service.generate_task_work_notes(
                                narrative_type=NarrativeType.TASK_FAILURE,
                                task_number=task_number,
                                short_description=task.get("short_description", ""),
                                description=task.get("description", ""),
                                technical_details=str(e),
                            )
                            logger.info(
                                "Diagnostic: updating work notes for %s with generated content: %s",
                                task_number,
                                notes,
                            )
                            update_task_work_notes(task_id, notes)
                        except Exception as n_err:
                            logger.error("Failed to generate/update failure narrative for %s: %s", task_number, n_err)

                    update_task_state(
                        task_id,
                        TASK_STATES["CLOSED_INCOMPLETE"],
                    )
            else:
                # Legacy Backward Compatibility Fallback Code path
                try:
                    result = process_task(task)
                    status = result.get("status")

                    final_state = (
                        TASK_STATES["CLOSED_COMPLETE"]
                        if status == "success"
                        else TASK_STATES["CLOSED_INCOMPLETE"]
                    )

                    update_task_state(task_id, final_state)
                    logger.info("%s completed synchronously with status=%s", task_number, status)

                except Exception as e:
                    logger.exception(
                        "Synchronous legacy execution failed for %s: %s",
                        task_number,
                        e,
                    )
                    update_task_state(
                        task_id,
                        TASK_STATES["CLOSED_INCOMPLETE"],
                    )

        except Exception as e:
            logger.exception("Poller error: %s", e)
            if not work_done:
                _wait(stop_event, int(POLL_INTERVAL) if POLL_INTERVAL is not None else 60)


def _track_failed_task(task_id: str, sctask: str, task: Dict[str, Any], failure_reason: str):
    """Track failed tasks (validation failed before CR creation) for dashboard visibility."""
    failed_tasks_file = "data/tracking/failed_tasks.json"
    
    try:
        # Load existing failed tasks
        if os.path.exists(failed_tasks_file):
            try:
                with open(failed_tasks_file, "r") as f:
                    data = json.load(f)
                    failed_tasks = data.get("failed_tasks", {})
            except Exception:
                failed_tasks = {}
        else:
            failed_tasks = {}
        
        # Add this failed task
        failed_tasks[task_id] = {
            "task_number": sctask,
            "short_description": task.get("short_description", ""),
            "device_name": task.get("device_data", {}).get("device_name", "Unknown"),
            "attempted_at": task.get("created_on", ""),
            "failed_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "failure_reason": failure_reason,
        }
        
        # Write back
        os.makedirs(os.path.dirname(failed_tasks_file), exist_ok=True)
        with open(failed_tasks_file, "w") as f:
            json.dump({"failed_tasks": failed_tasks}, f, indent=2)
        
        logger.info("Tracked failed task %s in %s", sctask, failed_tasks_file)
    except Exception as e:
        logger.error("Failed to track failed task %s: %s", sctask, e)


def _wait(stop_event, seconds: int | None):
    """Sleep that respects an optional stop_event."""
    if stop_event is None:
        time.sleep(seconds)
    else:
        stop_event.wait(seconds)
